import openpyxl
import pandas
from Phong import *
import os
cur_dir = os.path.abspath(".")
def reloadDB_ROOM(lst):
    wb = openpyxl.load_workbook(cur_dir + "/data/data.xlsx")
    ws = wb["phong"]
    t = 1
    
    for i in lst:
        t += 1
        ws[f"A{t}"].value = i.roomId
        ws[f"B{t}"].value = i.square
        ws[f"C{t}"].value = i.status
        ws[f"D{t}"].value = i.maxPeople
    t = 1
    for i in lst:
        t+=1
        if len(i.getlstCus()) == 0:
            i.status = "trong"
            ws[f"C{t}"].value = i.status
        else:
            i.status = "co_nguoi"
            ws[f"C{t}"].value = i.status
    t += 1
    ws[f"A{t}"].value = ''
    ws[f"B{t}"].value = ''
    ws[f"C{t}"].value = ''
    ws[f"D{t}"].value = ''
    wb.save(cur_dir + "/data/data.xlsx")
