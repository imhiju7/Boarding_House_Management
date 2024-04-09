import openpyxl
import pandas
# 
import os
cur_dir = os.path.abspath(".")
file= pandas.read_excel((cur_dir + "/data/data.xlsx"),sheet_name='khach') 
file=file.fillna("")
mylist=file.values.tolist()
def getlst():
    return mylist
def find():
    wb=openpyxl.load_workbook(cur_dir + "/data/data.xlsx")
    ws=wb["khach"]
    n=str(input("Nhập tên :"))
    for row in range(1, ws.max_row + 1):
        for column in "A":
            cell_name = "{}{}".format(column, row)
            if n in ws[cell_name].value:
                print("\n")
                for col in "ABCDEFG":
                    next_cell = "{}{}".format(col, row)
                    print(" ",ws[next_cell].value,end='')
def checkinfo(id:str,name:str,phone:str):
    if id.startswith("KH") == False:
        return False
    if name.strip() == "":
        return False
    if phone.isdigit() == False or len(phone) < 9 or len(phone) > 10:
        return False
def newExcel(lst):
    wb=openpyxl.load_workbook(cur_dir + "/data/data.xlsx")
    ws=wb["khach"]
    t = 1
    for i in lst:
        t += 1
        ws[f'A{t}'].value = i[0]
        ws[f'B{t}'].value = i[1]
        ws[f'C{t}'].value = i[2]
        ws[f'D{t}'].value = i[3]
        ws[f'E{t}'].value = i[4]
        ws[f'F{t}'].value = i[5]
        ws[f'G{t}'].value = i[6]
    t += 1
    ws[f'A{t}'].value = ''
    ws[f'B{t}'].value = ''
    ws[f'C{t}'].value = ''
    ws[f'D{t}'].value = ''
    ws[f'E{t}'].value = ''
    ws[f'F{t}'].value = ''
    ws[f'G{t}'].value = ''
    wb.save(cur_dir + "/data/data.xlsx")
    mylist = lst
